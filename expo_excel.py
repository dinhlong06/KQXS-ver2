import os
import openpyxl
import win32com.client

file = "example.xlsx"

PROVINCE_CONFIG = {
    "urls": {
        0: ["tp-hcm", "dong-thap", "ca-mau"],
        1: ["ben-tre", "vung-tau", "bac-lieu"],
        2: ["dong-nai", "can-tho", "soc-trang"],
        3: ["tay-ninh", "an-giang", "binh-thuan"],
        4: ["vinh-long", "binh-duong", "tra-vinh"],
        5: ["tp-hcm", "long-an", "binh-phuoc"],
        6: ["tien-giang", "kien-giang", "da-lat"]
    },
    "display_names": [
        ["TP.HCM", "Đ.THÁP", "CÀ MAU"],
        ["BẾN TRE", "V.TÀU", "BẠC LIÊU"],
        ["ĐỒNG NAI", "CẦN THƠ", "S.TRĂNG"],
        ["TÂY NINH", "AN GIANG", "B.THUẬN"],
        ["V.LONG", "B.DƯƠNG", "T.VINH"],
        ["TP.HCM", "LONG AN", "B.PHƯỚC"],
        ["T.GIANG", "K.GIANG", "ĐÀ LẠT"]
    ]
}

def close_excel_file(file_path):
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        for workbook in excel.Workbooks:
            if workbook.FullName == os.path.abspath(file_path):
                workbook.Close(SaveChanges=True)
        excel.Quit()
        return True
    except Exception as e:
        print(f"Không thể đóng file Excel: {e}")
        return False


def open_excel_file(file_path):
    try:
        os.startfile(file_path)
        return True
    except Exception as e:
        print(f"Không thể mở file Excel: {e}")
        return False

def clear_excel_range(file_path, sheet_name=None):
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        wb = excel.Workbooks.Open(os.path.abspath(file_path))
        ws = wb.Sheets(sheet_name) if sheet_name else wb.Sheets(1)
        ws.Range("B2:D19").ClearContents()
        wb.Save()
        wb.Close()
        excel.Quit()
        print("Đã clear vùng dữ liệu B2:D19 thành công.")
    except Exception as e:
        print(f"Lỗi khi clear vùng dữ liệu: {e}")

def write_to_excel(tinhs, weekday, ngayxoso, file_path=file):
    try:
        if not os.path.exists(file_path):
            wb = openpyxl.Workbook()
            wb.save(file_path)
            wb.close()

        # Clear sẵn vùng B2:D19
        clear_excel_range(file_path)

        # Load file và ghi dữ liệu
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        province_names = [
            ["TP.HCM", "Đ.THÁP", "CÀ MAU"],
            ["BẾN TRE", "V.TÀU", "BẠC LIÊU"],
            ["ĐỒNG NAI", "CẦN THƠ", "S.TRĂNG"],
            ["TÂY NINH", "AN GIANG", "B.THUẬN"],
            ["V.LONG", "B.DƯƠNG", "T.VINH"],
            ["TP.HCM", "LONG AN", "B.PHƯỚC"],
            ["T.GIANG", "K.GIANG", "ĐÀ LẠT"]
        ]

        for i in range(3):
            sheet.cell(1, i + 2).value = province_names[weekday][i]

        for i, tinh in enumerate(tinhs):
            if tinh:
                row = 2
                for key, values in tinh.items():
                    for value in values:
                        sheet.cell(row, i + 2).value = value
                        row += 1

        sheet.cell(24, 1).value = "KQXS NGÀY " + ngayxoso

        wb.save(file_path)
        wb.close()
        print("Ghi dữ liệu vào Excel thành công.")
    except Exception as e:
        print(f"Lỗi khi ghi vào Excel: {e}")